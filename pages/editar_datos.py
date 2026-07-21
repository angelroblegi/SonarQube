import os
import glob
import re

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from auth_utils import requiere_admin

st.set_page_config(layout="wide", page_title="Editar datos de componentes")

# Solo administradores pueden editar datos
requiere_admin()

UPLOAD_DIR = "uploads"
ARCHIVO_SELECCION = "data/seleccion_proyectos.csv"

# Columnas de métricas que se pueden editar
RATING_COLS = ["security_rating", "reliability_rating", "sqale_rating"]
METRIC_COLS = [
    "security_rating", "reliability_rating", "sqale_rating", "coverage",
    "duplicated_lines_density", "bugs", "bugs_blocker", "bugs_critical",
    "bugs_major", "bugs_minor", "bugs_info",
]
OPCIONES_RATING = ["A", "B", "C", "D", "E"]

st.title("🛠️ Editar datos de componentes")
st.caption(
    "Edita las métricas de un componente para un mes específico, "
    "o cambia la célula de un componente a través de varios meses."
)


# ---------------------- Utilidades de archivos ----------------------

def mes_de_archivo(path):
    """Extrae 'YYYY-MM' del nombre metricas_YYYY-MM.xlsx."""
    m = re.search(r"metricas_(\d{4}-\d{2})\.xlsx$", os.path.basename(path))
    return m.group(1) if m else None


def archivos_por_mes():
    """Devuelve un dict {mes: ruta} ordenado por mes."""
    archivos = glob.glob(os.path.join(UPLOAD_DIR, "metricas_*.xlsx"))
    mapa = {}
    for a in archivos:
        mes = mes_de_archivo(a)
        if mes:
            mapa[mes] = a
    return dict(sorted(mapa.items()))


def leer_headers(ws):
    """Devuelve {nombre_columna: indice_1based} usando la primera fila."""
    headers = {}
    for i, celda in enumerate(ws[1], start=1):
        if celda.value is not None:
            headers[str(celda.value).strip()] = i
    return headers


def nombre_col_mes(headers):
    for candidato in ("Mes", "mes"):
        if candidato in headers:
            return candidato
    return None


def convertir_valor(valor_str):
    """Convierte un texto a número si aplica; si no, deja el texto tal cual.
    Vacío -> None (celda vacía)."""
    if valor_str is None:
        return None
    s = str(valor_str).strip()
    if s == "":
        return None
    # Intentar entero
    try:
        if re.fullmatch(r"-?\d+", s):
            return int(s)
    except ValueError:
        pass
    # Intentar decimal (acepta coma o punto)
    try:
        return float(s.replace(",", "."))
    except ValueError:
        return s


@st.cache_data
def cargar_dataframe(path, _mtime):
    """Carga un archivo mensual como DataFrame (para mostrar/seleccionar).
    _mtime fuerza recarga cuando el archivo cambia."""
    df = pd.read_excel(path)
    df.columns = df.columns.str.strip()
    return df


def cargar_mes(path):
    return cargar_dataframe(path, os.path.getmtime(path))


def filas_de_proyecto(ws, headers, nombre_proyecto, celula=None):
    """Devuelve los índices (1-based) de filas cuyo NombreProyecto coincide."""
    np_idx = headers.get("NombreProyecto")
    cel_idx = headers.get("Celula")
    filas = []
    if np_idx is None:
        return filas
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=np_idx).value
        if val is not None and str(val).strip() == str(nombre_proyecto).strip():
            if celula is not None and cel_idx is not None:
                cel_val = ws.cell(row=r, column=cel_idx).value
                if str(cel_val).strip() != str(celula).strip():
                    continue
            filas.append(r)
    return filas


mapa_archivos = archivos_por_mes()

if not mapa_archivos:
    st.warning("⚠️ No se encontró ningún archivo de métricas en la carpeta 'uploads'.")
    st.stop()

tab_metricas, tab_celula = st.tabs([
    "✏️ Editar métricas (un mes)",
    "🔀 Cambiar célula (varios meses)",
])


# ============================================================
# FEATURE 1: Editar métricas de un componente en un mes
# ============================================================
with tab_metricas:
    st.subheader("Editar métricas de un componente")
    st.caption("Los cambios afectan únicamente al mes seleccionado.")

    meses = list(mapa_archivos.keys())
    mes_sel = st.selectbox(
        "1️⃣ Selecciona el mes",
        options=meses,
        index=len(meses) - 1,
        key="mes_metricas",
    )
    path_mes = mapa_archivos[mes_sel]
    df_mes = cargar_mes(path_mes)

    proyectos = sorted(df_mes["NombreProyecto"].dropna().astype(str).unique())
    proyecto_sel = st.selectbox(
        "2️⃣ Selecciona el componente",
        options=proyectos,
        index=None,
        placeholder="Escribe para buscar...",
        key="proyecto_metricas",
    )

    if proyecto_sel:
        fila_df = df_mes[df_mes["NombreProyecto"].astype(str) == proyecto_sel]

        if len(fila_df) > 1:
            st.info(
                f"ℹ️ Este componente aparece {len(fila_df)} veces en {mes_sel}. "
                "Se editarán todas sus filas con los mismos valores."
            )

        fila = fila_df.iloc[0]
        celula_actual = fila.get("Celula", "")
        st.markdown(f"**Célula actual:** `{celula_actual}`")

        # Columnas de métricas presentes en este archivo
        cols_presentes = [c for c in METRIC_COLS if c in df_mes.columns]

        with st.form("form_metricas"):
            nuevos_valores = {}
            columnas = st.columns(3)
            for i, col in enumerate(cols_presentes):
                valor_actual = fila.get(col, "")
                valor_actual = "" if pd.isna(valor_actual) else str(valor_actual)
                with columnas[i % 3]:
                    if col in RATING_COLS:
                        opciones = list(dict.fromkeys(
                            ([valor_actual] if valor_actual else []) + OPCIONES_RATING
                        ))
                        idx = opciones.index(valor_actual) if valor_actual in opciones else 0
                        nuevos_valores[col] = st.selectbox(col, opciones, index=idx)
                    else:
                        nuevos_valores[col] = st.text_input(col, value=valor_actual)

            guardar = st.form_submit_button("💾 Guardar cambios de métricas")

        if guardar:
            wb = load_workbook(path_mes)
            ws = wb.active
            headers = leer_headers(ws)
            filas = filas_de_proyecto(ws, headers, proyecto_sel)

            if not filas:
                st.error("No se encontró el componente en el archivo. No se guardó nada.")
            else:
                for r in filas:
                    for col, valor in nuevos_valores.items():
                        if col in headers:
                            ws.cell(row=r, column=headers[col]).value = convertir_valor(valor)
                wb.save(path_mes)
                st.cache_data.clear()
                st.success(
                    f"✅ Métricas actualizadas para '{proyecto_sel}' en {mes_sel} "
                    f"({len(filas)} fila(s))."
                )
                st.rerun()


# ============================================================
# FEATURE 2: Cambiar la célula de un componente en varios meses
# ============================================================
with tab_celula:
    st.subheader("Cambiar la célula de un componente")
    st.caption(
        "La célula está ligada al componente a lo largo de los meses. "
        "Aquí puedes reasignarlo en el rango de meses que elijas."
    )

    # Índice global de proyecto -> meses en los que aparece y su célula
    @st.cache_data
    def indice_global(firmas):
        registros = []
        for mes, path in mapa_archivos.items():
            df = cargar_mes(path)
            if "NombreProyecto" not in df.columns:
                continue
            for _, row in df[["NombreProyecto", "Celula"]].dropna(subset=["NombreProyecto"]).iterrows():
                registros.append({
                    "Mes": mes,
                    "NombreProyecto": str(row["NombreProyecto"]).strip(),
                    "Celula": "" if pd.isna(row.get("Celula")) else str(row["Celula"]).strip(),
                })
        return pd.DataFrame(registros)

    firmas = tuple((m, os.path.getmtime(p)) for m, p in mapa_archivos.items())
    df_global = indice_global(firmas)

    todos_proyectos = sorted(df_global["NombreProyecto"].unique())
    proyecto_cel = st.selectbox(
        "1️⃣ Selecciona el componente",
        options=todos_proyectos,
        index=None,
        placeholder="Escribe para buscar...",
        key="proyecto_celula",
    )

    if proyecto_cel:
        apariciones = df_global[df_global["NombreProyecto"] == proyecto_cel].sort_values("Mes")
        st.markdown("**Apariciones actuales:**")
        st.dataframe(apariciones[["Mes", "Celula"]], use_container_width=True, hide_index=True)

        meses_disponibles = apariciones["Mes"].tolist()
        celulas_existentes = sorted(df_global["Celula"].replace("", pd.NA).dropna().unique())

        st.markdown("**2️⃣ Selecciona los meses a actualizar**")
        aplicar_todos = st.checkbox(
            "Aplicar a todos los meses en los que aparece el componente",
            value=True,
            key="aplicar_todos",
        )
        if aplicar_todos:
            meses_objetivo = meses_disponibles
        else:
            meses_objetivo = st.multiselect(
                "Meses a actualizar",
                options=meses_disponibles,
                default=meses_disponibles,
                key="meses_objetivo",
            )

        st.markdown("**3️⃣ Nueva célula**")
        col_a, col_b = st.columns(2)
        with col_a:
            celula_elegida = st.selectbox(
                "Elegir una célula existente",
                options=["— (escribir una nueva) —"] + celulas_existentes,
                key="celula_existente",
            )
        with col_b:
            celula_nueva_txt = st.text_input(
                "…o escribir una nueva célula",
                key="celula_nueva_txt",
            )

        nueva_celula = (
            celula_nueva_txt.strip()
            if celula_nueva_txt.strip()
            else (celula_elegida if celula_elegida != "— (escribir una nueva) —" else "")
        )

        actualizar_seleccion = st.checkbox(
            "También actualizar la selección de proyectos (seleccion_proyectos.csv)",
            value=True,
            key="actualizar_seleccion",
        )

        if st.button("💾 Cambiar célula", key="btn_cambiar_celula"):
            if not nueva_celula:
                st.error("Debes elegir o escribir una nueva célula.")
            elif not meses_objetivo:
                st.error("Debes seleccionar al menos un mes.")
            else:
                total_filas = 0
                meses_editados = []
                for mes in meses_objetivo:
                    path = mapa_archivos[mes]
                    wb = load_workbook(path)
                    ws = wb.active
                    headers = leer_headers(ws)
                    if "Celula" not in headers:
                        continue
                    filas = filas_de_proyecto(ws, headers, proyecto_cel)
                    for r in filas:
                        ws.cell(row=r, column=headers["Celula"]).value = nueva_celula
                    if filas:
                        wb.save(path)
                        total_filas += len(filas)
                        meses_editados.append(mes)

                # Actualizar CSV de selección de proyectos si aplica
                sel_msg = ""
                if actualizar_seleccion and os.path.exists(ARCHIVO_SELECCION):
                    df_sel = pd.read_csv(ARCHIVO_SELECCION)
                    mask = df_sel["NombreProyecto"].astype(str) == proyecto_cel
                    if mask.any():
                        df_sel.loc[mask, "Celula"] = nueva_celula
                        df_sel = df_sel.drop_duplicates(subset=["Celula", "NombreProyecto"])
                        df_sel.to_csv(ARCHIVO_SELECCION, index=False)
                        sel_msg = " Selección de proyectos actualizada."

                st.cache_data.clear()
                if total_filas:
                    st.success(
                        f"✅ Célula de '{proyecto_cel}' cambiada a '{nueva_celula}' "
                        f"en {len(meses_editados)} mes(es): {', '.join(meses_editados)}."
                        + sel_msg
                    )
                    st.rerun()
                else:
                    st.warning("No se actualizó ninguna fila (no se encontró el componente en los meses elegidos).")
